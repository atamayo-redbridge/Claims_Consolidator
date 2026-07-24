from pathlib import Path
import textwrap, zipfile, os, json

base = Path("/mnt/data/RedbridgeLargeClaims")
base.mkdir(parents=True, exist_ok=True)

files = {}

files["requirements.txt"] = """\
streamlit>=1.36
pandas>=2.2
openpyxl>=3.1
xlrd>=2.0
"""

files["contracts.py"] = r'''from __future__ import annotations

from copy import deepcopy
from typing import Any


# IMPORTANT:
# - Keys are Group Number -> Policy Year.
# - Monetary amounts are stored as numbers, without currency symbols.
# - A member rule can be:
#       {"type": "laser", "deductible": 250000}
#       {"type": "excluded"}
# - Some contract details were not supplied. Those entries remain None and
#   the application will warn the user instead of silently guessing.


CONTRACTS: dict[str, dict[str, dict[str, Any]]] = {
    "750133": {
        "2024": {
            "company": "International Hospitality Services",
            "deductible": 100000,
            "maximum_liability": 900000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23389637700": {"type": "laser", "deductible": 250000},
                "23398547100": {"type": "laser", "deductible": 150000},
            },
        },
        "2025": {
            "company": "International Hospitality Services",
            "deductible": 100000,
            "maximum_liability": 1000000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23389637700": {"type": "laser", "deductible": 300000},
                "23400968000": {"type": "laser", "deductible": 200000},
                "23389606301": {"type": "laser", "deductible": 150000},
            },
        },
    },
    "750134": {
        "2024": {
            "company": "CONWASTE",
            "deductible": 150000,
            "maximum_liability": 850000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {},
        },
        "2025": {
            "company": "CONWASTE",
            "deductible": 150000,
            "maximum_liability": 850000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {},
        },
    },
    "750102": {
        "2025": {
            "company": "Cardinal Health PR",
            "deductible": 150000,
            "maximum_liability": 850000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23325932000": {"type": "laser", "deductible": 500000},
                "23325936302": {"type": "laser", "deductible": 200000},
                "23325938801": {"type": "laser", "deductible": 300000},
                "23321395901": {"type": "laser", "deductible": 225000},
                "23325914301": {"type": "laser", "deductible": 200000},
            },
        },
        "2026": {
            "company": "Cardinal Health PR",
            "deductible": None,  # Pending confirmation
            "maximum_liability": None,  # Pending confirmation
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23325938801": {"type": "laser", "deductible": 350000},
                "23395124702": {"type": "laser", "deductible": 360000},
                "23395114800": {"type": "laser", "deductible": 250000},
            },
        },
    },
    "750136": {
        "2025": {
            "company": "FENWAL INTERNATIONAL INC",
            "deductible": 125000,
            "maximum_liability": 875000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {},
            "notes": "A 2025 laser exists, but the Member ID and amount are still pending.",
        },
        "2026": {
            "company": "FENWAL INTERNATIONAL INC",
            "deductible": None,  # Pending confirmation
            "maximum_liability": None,  # Pending confirmation
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23414805502": {"type": "excluded"},
            },
        },
    },
    "750132": {
        "2024": {
            "company": "Grupo Cooperativo Seguros Multiples",
            "deductible": 100000,
            "maximum_liability": 900000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23382390001": {"type": "excluded"},
                "23382407600": {"type": "laser", "deductible": 150000},
            },
        },
        "2025": {
            "company": "Grupo Cooperativo Seguros Multiples",
            "deductible": 100000,
            "maximum_liability": 900000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23382390001": {"type": "excluded"},
                "23382407600": {"type": "laser", "deductible": 150000},
            },
        },
    },
    "750123": {},
    "750124": {},
    "750125": {},
    "711205": {},
    "750101": {
        "2025": {
            "company": "National University College",
            "deductible": 100000,
            "maximum_liability": 900000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {},
        },
    },
    "750109": {
        "2025": {
            "company": "Oriental Bank",
            "deductible": 175000,
            "maximum_liability": 1000000,
            "covered_benefits": ["MED", "RX", "DENTAL"],
            "member_rules": {
                "23397560700": {"type": "laser", "deductible": 500000},
                "23397485400": {"type": "laser", "deductible": 500000},
                "23397502801": {"type": "laser", "deductible": 250000},
                "23397483300": {"type": "laser", "deductible": 250000},
            },
        },
    },
    "750093": {
        "2024": {
            "company": "Universal Group",
            "deductible": 150000,
            "maximum_liability": 1000000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23383276501": {"type": "excluded"},
            },
        },
        "2025": {
            "company": "Universal Group",
            "deductible": 200000,
            "maximum_liability": 2000000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {
                "23390024204": {"type": "laser", "deductible": 200000},
                "23276817500": {"type": "laser", "deductible": 200000},
                "23408941000": {"type": "laser", "deductible": 500000},
                "23383276501": {"type": "excluded"},
            },
        },
    },
    "750096": {
        "2026": {
            "company": "Walmart",
            "deductible": 125000,
            "maximum_liability": 875000,
            "covered_benefits": ["MED", "RX"],
            "member_rules": {},
        },
    },
}


LIBERTY_YEARS = {
    "2024": {
        "company": "Liberty Communications of PR",
        "deductible": 200000,
        "maximum_liability": 800000,
        "covered_benefits": ["MED", "RX"],
        "member_rules": {
            "23409445301": {"type": "laser", "deductible": 350000},
            "23409212300": {"type": "laser", "deductible": 350000},
            "23283558500": {"type": "laser", "deductible": 350000},
        },
    },
    "2025": {
        "company": "Liberty Communications of PR",
        "deductible": 200000,
        "maximum_liability": 800000,
        "covered_benefits": ["MED", "RX"],
        "member_rules": {
            "23277353200": {"type": "excluded"},
            "23270105604": {"type": "excluded"},
        },
    },
    "2026": {
        "company": "Liberty Communications of PR",
        "deductible": 200000,
        "maximum_liability": 800000,
        "covered_benefits": ["MED", "RX"],
        "member_rules": {
            "23270105604": {"type": "laser", "deductible": 400000},
            "23378173801": {"type": "laser", "deductible": 250000},
            "23282225204": {"type": "laser", "deductible": 300000},
            "23409521000": {"type": "laser", "deductible": 250000},
            "23277673401": {"type": "laser", "deductible": 300000},
            "23270857301": {"type": "laser", "deductible": 300000},
        },
    },
}

for liberty_group in ("750123", "750124", "750125", "711205"):
    CONTRACTS[liberty_group] = deepcopy(LIBERTY_YEARS)


def get_contract(group_number: str, policy_year: str) -> dict[str, Any]:
    group_number = str(group_number).strip()
    policy_year = str(policy_year).strip()

    if group_number not in CONTRACTS:
        raise KeyError(f"Group {group_number} is not in the contract catalog.")

    if policy_year not in CONTRACTS[group_number]:
        available = ", ".join(sorted(CONTRACTS[group_number].keys())) or "none"
        raise KeyError(
            f"Policy year {policy_year} is not configured for group {group_number}. "
            f"Available years: {available}."
        )

    contract = deepcopy(CONTRACTS[group_number][policy_year])
    contract["group_number"] = group_number
    contract["policy_year"] = policy_year
    return contract
'''

files["claims_reader.py"] = r'''from __future__ import annotations

import io
import re
from pathlib import Path
from typing import BinaryIO

import pandas as pd


GROUP_CANDIDATES = [
    "GROUP NUMBER", "GROUP NUMBER ", "GRPNUM", "GROUP", "GROUP NO",
    "GROUP #", "COMPNO", "COMP NO", "COMPANY NUMBER"
]

MEMBER_CANDIDATES = [
    "MEMBER ID NUMBER", "MEMBER ID", "MEMBERID", "MEMBNO",
    "MEMBER NUMBER", "MEMBER NO", "CERTIFICATE NUMBER"
]

AMOUNT_CANDIDATES = [
    "AMOUNT PAID", "PAID AMOUNT", "COMPUTED", "CLAIM AMOUNT",
    "NET PAID", "TOTAL PAID", "AMOUNT", "PAID"
]

FIRST_NAME_CANDIDATES = [
    "MEMB FIRST NAME", "MEMBER FIRST NAME", "FIRST NAME", "FSTNAM",
    "FIRST"
]

LAST_NAME_CANDIDATES = [
    "MEMB LAS NAME", "MEMB LAST NAME", "MEMBER LAST NAME",
    "LAST NAME", "LSTNAM", "LAST"
]


def _clean_column_name(value: object) -> str:
    text = str(value).strip().upper()
    text = re.sub(r"\s+", " ", text)
    return text


def _find_column(columns: list[str], candidates: list[str]) -> str | None:
    normalized = {_clean_column_name(col): col for col in columns}
    for candidate in candidates:
        if _clean_column_name(candidate) in normalized:
            return normalized[_clean_column_name(candidate)]

    # Conservative partial matching fallback.
    for normalized_name, original in normalized.items():
        for candidate in candidates:
            candidate_norm = _clean_column_name(candidate)
            if candidate_norm in normalized_name or normalized_name in candidate_norm:
                return original
    return None


def normalize_identifier(value: object, append_00_to_9_digits: bool = False) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip()

    # Excel frequently converts IDs to floats such as 234045521.0.
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".")[0]

    digits = re.sub(r"\D", "", text)

    if append_00_to_9_digits and len(digits) == 9:
        digits += "00"

    return digits


def _read_data_sheet(uploaded_file: BinaryIO) -> pd.DataFrame:
    raw = uploaded_file.getvalue() if hasattr(uploaded_file, "getvalue") else uploaded_file.read()
    buffer = io.BytesIO(raw)

    excel = pd.ExcelFile(buffer)
    matching_sheet = next(
        (sheet for sheet in excel.sheet_names if sheet.strip().lower() == "data"),
        None,
    )
    if matching_sheet is None:
        raise ValueError(
            f"The workbook does not contain a sheet named 'data'. "
            f"Available sheets: {', '.join(excel.sheet_names)}"
        )

    return pd.read_excel(excel, sheet_name=matching_sheet)


def infer_benefit_type(filename: str, columns: list[str]) -> str:
    name = Path(filename).stem.upper()
    joined_columns = " ".join(_clean_column_name(c) for c in columns)

    if "DENTAL" in name or "DENTAL" in joined_columns:
        return "DENTAL"
    if "VISION" in name or "VISION" in joined_columns:
        return "VISION"
    if re.search(r"(^|[^A-Z])RX([^A-Z]|$)", name) or "PRESCRIPTION" in joined_columns:
        return "RX"
    if "PHARM" in name:
        return "RX"
    return "MED"


def read_claim_file(uploaded_file: BinaryIO) -> pd.DataFrame:
    df = _read_data_sheet(uploaded_file)
    if df.empty:
        raise ValueError(f"{uploaded_file.name}: the 'data' sheet is empty.")

    columns = list(df.columns)
    group_col = _find_column(columns, GROUP_CANDIDATES)
    member_col = _find_column(columns, MEMBER_CANDIDATES)
    amount_col = _find_column(columns, AMOUNT_CANDIDATES)
    first_col = _find_column(columns, FIRST_NAME_CANDIDATES)
    last_col = _find_column(columns, LAST_NAME_CANDIDATES)

    missing = []
    if not group_col:
        missing.append("Group Number")
    if not member_col:
        missing.append("Member ID")
    if not amount_col:
        missing.append("Claim Amount")
    if missing:
        raise ValueError(
            f"{uploaded_file.name}: could not detect {', '.join(missing)}. "
            f"Columns found: {', '.join(map(str, columns))}"
        )

    benefit = infer_benefit_type(uploaded_file.name, columns)

    out = pd.DataFrame()
    out["Group Number"] = df[group_col].map(normalize_identifier)
    out["Member ID"] = df[member_col].map(
        lambda x: normalize_identifier(x, append_00_to_9_digits=True)
    )
    out["First Name"] = (
        df[first_col].fillna("").astype(str).str.strip() if first_col else ""
    )
    out["Last Name"] = (
        df[last_col].fillna("").astype(str).str.strip() if last_col else ""
    )
    out["Benefit"] = benefit
    out["Claim Amount"] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)
    out["Source File"] = uploaded_file.name

    out = out[
        (out["Group Number"] != "")
        & (out["Member ID"] != "")
        & (out["Claim Amount"] != 0)
    ].copy()

    if out.empty:
        raise ValueError(
            f"{uploaded_file.name}: no usable claim rows remained after cleaning."
        )

    return out


def read_all_claim_files(uploaded_files: list[BinaryIO]) -> pd.DataFrame:
    if not uploaded_files:
        raise ValueError("Upload at least one claim workbook.")

    frames = [read_claim_file(file) for file in uploaded_files]
    combined = pd.concat(frames, ignore_index=True)

    groups = sorted(combined["Group Number"].dropna().unique().tolist())
    if len(groups) != 1:
        raise ValueError(
            "The uploaded files must contain exactly one Group Number. "
            f"Detected groups: {', '.join(groups)}"
        )

    return combined
'''

files["claims_engine.py"] = r'''from __future__ import annotations

from typing import Any

import pandas as pd


def consolidate_claims(raw_claims: pd.DataFrame) -> pd.DataFrame:
    """Return one row per Group Number + Member ID."""
    if raw_claims.empty:
        raise ValueError("No claim records were supplied.")

    name_lookup = (
        raw_claims.sort_values(["Member ID", "First Name", "Last Name"])
        .groupby(["Group Number", "Member ID"], as_index=False)
        .agg(
            {
                "First Name": lambda s: next((v for v in s if str(v).strip()), ""),
                "Last Name": lambda s: next((v for v in s if str(v).strip()), ""),
            }
        )
    )

    benefit_totals = (
        raw_claims.pivot_table(
            index=["Group Number", "Member ID"],
            columns="Benefit",
            values="Claim Amount",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reset_index()
    )
    benefit_totals.columns.name = None

    result = name_lookup.merge(
        benefit_totals, on=["Group Number", "Member ID"], how="outer"
    )

    benefit_columns = [
        c for c in result.columns
        if c not in {"Group Number", "Member ID", "First Name", "Last Name"}
    ]

    for column in benefit_columns:
        result[column] = pd.to_numeric(result[column], errors="coerce").fillna(0.0)

    result["Total Claims"] = result[benefit_columns].sum(axis=1)
    return result.sort_values("Total Claims", ascending=False).reset_index(drop=True)


def apply_contract_rules(
    consolidated: pd.DataFrame,
    contract: dict[str, Any],
    alternative_deductible: float | None = None,
    replace_lasers: bool = False,
) -> pd.DataFrame:
    deductible = contract.get("deductible")
    maximum_liability = contract.get("maximum_liability")

    if deductible is None and alternative_deductible is None:
        raise ValueError(
            "The contract deductible is missing. Enter an Alternative Deductible "
            "before running the final analysis."
        )
    if maximum_liability is None:
        raise ValueError(
            "The contract maximum liability is missing from contracts.py. "
            "Add the confirmed amount before calculating Redbridge liability."
        )

    standard_deductible = (
        float(alternative_deductible)
        if alternative_deductible is not None
        else float(deductible)
    )

    member_rules = contract.get("member_rules", {})
    output = consolidated.copy()

    rule_types = []
    contract_lasers = []
    applicable_deductibles = []
    statuses = []
    liabilities = []
    above_limits = []

    for _, row in output.iterrows():
        member_id = str(row["Member ID"])
        total_claims = float(row["Total Claims"])
        rule = member_rules.get(member_id, {"type": "standard"})
        rule_type = rule.get("type", "standard").lower()

        contract_laser = None
        if rule_type == "excluded":
            applicable = None
            status = "Excluded - No Coverage"
            liability = 0.0
            above_limit = 0.0
        else:
            if rule_type == "laser":
                contract_laser = float(rule["deductible"])
                applicable = (
                    standard_deductible if replace_lasers else contract_laser
                )
                status = "Laser"
            else:
                applicable = standard_deductible
                status = "Standard"

            amount_above_deductible = max(total_claims - float(applicable), 0.0)
            liability = min(amount_above_deductible, float(maximum_liability))
            above_limit = max(
                amount_above_deductible - float(maximum_liability), 0.0
            )

        rule_types.append(rule_type.title())
        contract_lasers.append(contract_laser)
        applicable_deductibles.append(applicable)
        statuses.append(status)
        liabilities.append(liability)
        above_limits.append(above_limit)

    output["Rule Type"] = rule_types
    output["Contract Laser"] = contract_lasers
    output["Applicable Deductible"] = applicable_deductibles
    output["Coverage Status"] = statuses
    output["Redbridge Liability"] = liabilities
    output["Above Coverage Limit"] = above_limits
    output["Exceeds Deductible"] = output["Redbridge Liability"] > 0
    output["Policy Year"] = contract["policy_year"]
    output["Company"] = contract["company"]
    output["Maximum Redbridge Liability"] = float(maximum_liability)
    output["Analysis Deductible"] = standard_deductible
    output["Analysis Mode"] = (
        "Alternative deductible replaces all deductibles"
        if alternative_deductible is not None and replace_lasers
        else "Alternative deductible; contract lasers retained"
        if alternative_deductible is not None
        else "Contract terms"
    )

    preferred = [
        "Company", "Group Number", "Policy Year", "Member ID",
        "First Name", "Last Name",
    ]
    benefit_columns = [
        c for c in output.columns
        if c in {"MED", "RX", "DENTAL", "VISION"}
    ]
    ending = [
        "Total Claims", "Rule Type", "Contract Laser",
        "Applicable Deductible", "Coverage Status",
        "Maximum Redbridge Liability", "Redbridge Liability",
        "Above Coverage Limit", "Exceeds Deductible",
        "Analysis Deductible", "Analysis Mode",
    ]

    ordered = preferred + benefit_columns + ending
    remaining = [c for c in output.columns if c not in ordered]
    return output[ordered + remaining].sort_values(
        "Total Claims", ascending=False
    ).reset_index(drop=True)
'''

files["excel_report.py"] = r'''from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MONEY_COLUMNS = {
    "MED", "RX", "DENTAL", "VISION", "Total Claims", "Contract Laser",
    "Applicable Deductible", "Maximum Redbridge Liability",
    "Redbridge Liability", "Above Coverage Limit", "Analysis Deductible",
}


def _contract_summary(contract: dict[str, Any], analysis: pd.DataFrame) -> pd.DataFrame:
    lasers = []
    excluded = []

    for member_id, rule in contract.get("member_rules", {}).items():
        if rule.get("type") == "laser":
            lasers.append(f"{member_id}: ${rule['deductible']:,.2f}")
        elif rule.get("type") == "excluded":
            excluded.append(member_id)

    analysis_deductible = (
        analysis["Analysis Deductible"].iloc[0] if not analysis.empty else None
    )
    analysis_mode = analysis["Analysis Mode"].iloc[0] if not analysis.empty else ""

    rows = [
        ("Company", contract.get("company", "")),
        ("Group Number", contract.get("group_number", "")),
        ("Policy Year", contract.get("policy_year", "")),
        ("Contract Deductible", contract.get("deductible")),
        ("Analysis Deductible", analysis_deductible),
        ("Maximum Redbridge Liability", contract.get("maximum_liability")),
        ("Covered Benefits", ", ".join(contract.get("covered_benefits", []))),
        ("Analysis Mode", analysis_mode),
        ("Laser Members", "\n".join(lasers) if lasers else "None"),
        ("Excluded Members", "\n".join(excluded) if excluded else "None"),
        ("Notes", contract.get("notes", "")),
    ]
    return pd.DataFrame(rows, columns=["Field", "Value"])


def build_excel_report(
    analysis: pd.DataFrame,
    raw_claims: pd.DataFrame,
    contract: dict[str, Any],
) -> bytes:
    output = BytesIO()

    large_claims = analysis[analysis["Redbridge Liability"] > 0].copy()
    excluded = analysis[analysis["Coverage Status"] == "Excluded - No Coverage"].copy()
    summary = _contract_summary(contract, analysis)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        large_claims.to_excel(writer, sheet_name="Large Claims", index=False)
        analysis.to_excel(writer, sheet_name="All Members", index=False)
        excluded.to_excel(writer, sheet_name="Excluded Members", index=False)
        raw_claims.to_excel(writer, sheet_name="Raw Claims", index=False)

        workbook = writer.book

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            header_fill = PatternFill("solid", fgColor="1F4E78")
            for cell in worksheet[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for column_cells in worksheet.columns:
                letter = get_column_letter(column_cells[0].column)
                width = min(
                    max(len(str(cell.value)) if cell.value is not None else 0
                        for cell in column_cells) + 2,
                    40,
                )
                worksheet.column_dimensions[letter].width = max(width, 12)

            header_map = {
                cell.value: cell.column for cell in worksheet[1]
            }
            for money_column in MONEY_COLUMNS:
                column_index = header_map.get(money_column)
                if column_index:
                    for row in range(2, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=column_index).number_format = (
                            '$#,##0.00'
                        )

            if worksheet.title == "Summary":
                worksheet.column_dimensions["A"].width = 32
                worksheet.column_dimensions["B"].width = 65
                for row in worksheet.iter_rows(min_row=2, max_col=2):
                    row[1].alignment = Alignment(wrap_text=True, vertical="top")

    return output.getvalue()
'''

files["app.py"] = r'''from __future__ import annotations

import pandas as pd
import streamlit as st

from claims_engine import apply_contract_rules, consolidate_claims
from claims_reader import read_all_claim_files
from contracts import get_contract
from excel_report import build_excel_report


st.set_page_config(
    page_title="Redbridge Large Claims Analyzer",
    page_icon="📊",
    layout="wide",
)

st.title("Redbridge Large Claims Analyzer")
st.caption(
    "Upload claim workbooks, enter the policy year, consolidate claims by Member ID, "
    "and calculate Redbridge liability."
)

if "raw_claims" not in st.session_state:
    st.session_state.raw_claims = None
if "consolidated" not in st.session_state:
    st.session_state.consolidated = None
if "contract" not in st.session_state:
    st.session_state.contract = None
if "analysis" not in st.session_state:
    st.session_state.analysis = None


with st.container(border=True):
    uploaded_files = st.file_uploader(
        "Upload claim workbooks",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Each workbook must contain a sheet named 'data'.",
    )

    policy_year = st.text_input(
        "Policy Year",
        placeholder="Example: 2025",
        max_chars=4,
    )

    analyze_clicked = st.button("Analyze", type="primary", use_container_width=True)


if analyze_clicked:
    st.session_state.analysis = None

    try:
        year = policy_year.strip()
        if not year.isdigit() or len(year) != 4:
            raise ValueError("Enter a valid four-digit Policy Year.")

        raw_claims = read_all_claim_files(uploaded_files or [])
        group_number = str(raw_claims["Group Number"].iloc[0])
        contract = get_contract(group_number, year)
        consolidated = consolidate_claims(raw_claims)

        st.session_state.raw_claims = raw_claims
        st.session_state.consolidated = consolidated
        st.session_state.contract = contract

    except Exception as exc:
        st.session_state.raw_claims = None
        st.session_state.consolidated = None
        st.session_state.contract = None
        st.error(str(exc))


contract = st.session_state.contract

if contract:
    st.success("Contract found.")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Company", contract["company"])
    col2.metric("Group Number", contract["group_number"])
    col3.metric(
        "Contract Deductible",
        f"${contract['deductible']:,.2f}"
        if contract.get("deductible") is not None else "Pending",
    )
    col4.metric(
        "Maximum Liability",
        f"${contract['maximum_liability']:,.2f}"
        if contract.get("maximum_liability") is not None else "Pending",
    )

    st.subheader("Contract Information")
    st.write(
        "**Covered Benefits:** "
        + ", ".join(contract.get("covered_benefits", []))
    )

    uploaded_benefits = sorted(
        st.session_state.raw_claims["Benefit"].dropna().unique().tolist()
    )
    required_benefits = set(contract.get("covered_benefits", []))
    uploaded_set = set(uploaded_benefits)

    missing_benefits = sorted(required_benefits - uploaded_set)
    noncovered_uploaded = sorted(uploaded_set - required_benefits)

    if missing_benefits:
        st.warning(
            "Covered claim files not detected: " + ", ".join(missing_benefits)
        )
    else:
        st.info("All configured covered benefits were detected.")

    if noncovered_uploaded:
        st.warning(
            "Uploaded benefit types not covered by this contract: "
            + ", ".join(noncovered_uploaded)
            + ". They will be excluded from the contract analysis."
        )

    member_rules = contract.get("member_rules", {})
    laser_rows = []
    excluded_rows = []

    for member_id, rule in member_rules.items():
        if rule.get("type") == "laser":
            laser_rows.append(
                {
                    "Member ID": member_id,
                    "Laser Deductible": rule["deductible"],
                }
            )
        elif rule.get("type") == "excluded":
            excluded_rows.append(
                {
                    "Member ID": member_id,
                    "Rule": "No Coverage",
                }
            )

    left, right = st.columns(2)
    with left:
        st.markdown("#### Laser Members")
        if laser_rows:
            laser_df = pd.DataFrame(laser_rows)
            st.dataframe(
                laser_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Laser Deductible": st.column_config.NumberColumn(
                        format="$%.2f"
                    )
                },
            )
        else:
            st.write("None")

    with right:
        st.markdown("#### Excluded Members")
        if excluded_rows:
            st.dataframe(
                pd.DataFrame(excluded_rows),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.write("None")

    if contract.get("notes"):
        st.warning(contract["notes"])

    with st.container(border=True):
        st.subheader("Analysis Options")
        alternative_text = st.text_input(
            "Optional Alternative Deductible",
            placeholder="Leave blank to use the contract deductible",
            help="Enter numbers only, for example 75000.",
        )

        laser_mode = st.radio(
            "When an alternative deductible is entered:",
            options=[
                "Keep contract lasers",
                "Replace all deductibles, including lasers",
            ],
            horizontal=True,
        )

        run_final = st.button(
            "Run Final Analysis",
            type="primary",
            use_container_width=True,
        )

    if run_final:
        try:
            alternative = None
            if alternative_text.strip():
                cleaned = alternative_text.replace("$", "").replace(",", "").strip()
                alternative = float(cleaned)
                if alternative < 0:
                    raise ValueError(
                        "The Alternative Deductible cannot be negative."
                    )

            covered = set(contract.get("covered_benefits", []))
            filtered_raw = st.session_state.raw_claims[
                st.session_state.raw_claims["Benefit"].isin(covered)
            ].copy()

            if filtered_raw.empty:
                raise ValueError(
                    "None of the uploaded claims belong to a covered benefit."
                )

            filtered_consolidated = consolidate_claims(filtered_raw)
            analysis = apply_contract_rules(
                filtered_consolidated,
                contract,
                alternative_deductible=alternative,
                replace_lasers=(
                    laser_mode == "Replace all deductibles, including lasers"
                ),
            )
            st.session_state.analysis = analysis

        except Exception as exc:
            st.session_state.analysis = None
            st.error(str(exc))


analysis = st.session_state.analysis

if analysis is not None:
    st.divider()
    st.subheader("Analysis Results")

    total_claims = float(analysis["Total Claims"].sum())
    redbridge_liability = float(analysis["Redbridge Liability"].sum())
    large_claim_count = int((analysis["Redbridge Liability"] > 0).sum())
    excluded_count = int(
        (analysis["Coverage Status"] == "Excluded - No Coverage").sum()
    )

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Claims", f"${total_claims:,.2f}")
    c2.metric("Redbridge Liability", f"${redbridge_liability:,.2f}")
    c3.metric("Members Above Deductible", f"{large_claim_count:,}")
    c4.metric("Excluded Members Found", f"{excluded_count:,}")

    st.dataframe(
        analysis,
        use_container_width=True,
        hide_index=True,
        column_config={
            "MED": st.column_config.NumberColumn(format="$%.2f"),
            "RX": st.column_config.NumberColumn(format="$%.2f"),
            "DENTAL": st.column_config.NumberColumn(format="$%.2f"),
            "VISION": st.column_config.NumberColumn(format="$%.2f"),
            "Total Claims": st.column_config.NumberColumn(format="$%.2f"),
            "Contract Laser": st.column_config.NumberColumn(format="$%.2f"),
            "Applicable Deductible": st.column_config.NumberColumn(format="$%.2f"),
            "Maximum Redbridge Liability": st.column_config.NumberColumn(
                format="$%.2f"
            ),
            "Redbridge Liability": st.column_config.NumberColumn(format="$%.2f"),
            "Above Coverage Limit": st.column_config.NumberColumn(format="$%.2f"),
            "Exceeds Deductible": st.column_config.CheckboxColumn(),
        },
    )

    report_bytes = build_excel_report(
        analysis=analysis,
        raw_claims=st.session_state.raw_claims,
        contract=st.session_state.contract,
    )

    safe_company = (
        st.session_state.contract["company"]
        .replace(" ", "_")
        .replace("/", "_")
    )
    filename = (
        f"{safe_company}_{st.session_state.contract['group_number']}_"
        f"{st.session_state.contract['policy_year']}_Large_Claims.xlsx"
    )

    st.download_button(
        "Download Excel Report",
        data=report_bytes,
        file_name=filename,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        use_container_width=True,
    )
'''

files["README.md"] = r'''# Redbridge Large Claims Analyzer

## What it does

- Reads the `data` sheet from one or more Excel claim workbooks.
- Detects MED, RX, Dental, or Vision claim files.
- Normalizes Member IDs.
- Appends `00` when a Member ID contains exactly 9 digits.
- Sums all claims by Group Number and Member ID.
- Finds the contract using Group Number and the manually entered Policy Year.
- Displays company, deductible, maximum liability, covered benefits, lasers, and exclusions.
- Allows an optional alternative deductible for hypothetical analyses.
- Calculates Redbridge liability and amounts above the coverage limit.
- Produces a formatted Excel report.

## Project files

- `app.py`: Streamlit interface.
- `contracts.py`: contract catalog.
- `claims_reader.py`: Excel reader and column detection.
- `claims_engine.py`: claim consolidation and contract calculations.
- `excel_report.py`: Excel report creation.
- `requirements.txt`: Python packages.

## Installation

Open a terminal inside the project folder and run:

```bash
pip install -r requirements.txt


