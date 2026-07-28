from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


MONEY_COLUMNS = {
    "MED",
    "RX",
    "DENT",
    "DENTAL",  # Kept for compatibility with older analyses.
    "VISION",
    "Total Claims",
    "Contract Laser",
    "Applicable Deductible",
    "Maximum Redbridge Liability",
    "Redbridge Liability",
    "Above Coverage Limit",
    "Analysis Deductible",
}


def _normalize_benefit_name(value: object) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip().upper()

    aliases = {
        "DENTAL": "DENT",
        "DENT": "DENT",
        "MEDICAL": "MED",
        "MED": "MED",
        "PHARMACY": "RX",
        "PRESCRIPTION": "RX",
        "RX": "RX",
        "VISION": "VISION",
        "VIS": "VISION",
    }

    return aliases.get(text, text)


def _prepare_analysis(analysis: pd.DataFrame) -> pd.DataFrame:
    """
    Standardize the exported benefit columns.

    If an older analysis still contains DENTAL, it is renamed or merged
    into DENT so the final workbook uses the same naming convention as
    the current application.
    """
    prepared = analysis.copy()

    if "DENTAL" in prepared.columns and "DENT" not in prepared.columns:
        prepared = prepared.rename(
            columns={"DENTAL": "DENT"}
        )

    elif "DENTAL" in prepared.columns and "DENT" in prepared.columns:
        prepared["DENT"] = (
            pd.to_numeric(
                prepared["DENT"],
                errors="coerce",
            ).fillna(0.0)
            + pd.to_numeric(
                prepared["DENTAL"],
                errors="coerce",
            ).fillna(0.0)
        )

        prepared = prepared.drop(
            columns=["DENTAL"]
        )

    return prepared


def _prepare_raw_claims(raw_claims: pd.DataFrame) -> pd.DataFrame:
    prepared = raw_claims.copy()

    if "Benefit" in prepared.columns:
        prepared["Benefit"] = prepared[
            "Benefit"
        ].apply(_normalize_benefit_name)

    return prepared


def _style_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    headers = {
        cell.value: cell.column
        for cell in worksheet[1]
    }

    for column_name in MONEY_COLUMNS:
        if column_name in headers:
            column_index = headers[column_name]

            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):
                worksheet.cell(
                    row_number,
                    column_index,
                ).number_format = "$#,##0.00"

    for cells in worksheet.columns:
        column_letter = get_column_letter(
            cells[0].column
        )

        width = min(
            max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in cells
            )
            + 2,
            40,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = max(
            width,
            12,
        )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def build_excel_report(
    analysis: pd.DataFrame,
    raw_claims: pd.DataFrame,
    contract: dict,
) -> bytes:
    """
    Build one Excel report for one company/group.

    The updated app calls this function separately for each Group Number,
    so each downloaded workbook contains only that company's claims,
    contract terms, lasers, exclusions, and liability calculation.
    """
    prepared_analysis = _prepare_analysis(
        analysis
    )

    prepared_raw_claims = _prepare_raw_claims(
        raw_claims
    )

    lasers = []
    excluded = []

    for member, rule in contract.get(
        "member_rules",
        {},
    ).items():

        if rule.get("type") == "laser":
            lasers.append(
                f"{member}: "
                f"${rule['deductible']:,.2f}"
            )

        elif rule.get("type") == "excluded":
            excluded.append(
                member
            )

    covered_benefits = [
        _normalize_benefit_name(benefit)
        for benefit in contract.get(
            "covered_benefits",
            [],
        )
    ]

    summary = pd.DataFrame(
        [
            (
                "Company",
                contract.get("company", ""),
            ),
            (
                "Group Number",
                contract.get("group_number", ""),
            ),
            (
                "Policy Year",
                contract.get("policy_year", ""),
            ),
            (
                "Contract Deductible",
                contract.get("deductible"),
            ),
            (
                "Analysis Deductible",
                (
                    prepared_analysis[
                        "Analysis Deductible"
                    ].iloc[0]
                    if (
                        not prepared_analysis.empty
                        and "Analysis Deductible"
                        in prepared_analysis.columns
                    )
                    else None
                ),
            ),
            (
                "Maximum Redbridge Liability",
                contract.get(
                    "maximum_liability"
                ),
            ),
            (
                "Covered Benefits",
                ", ".join(
                    covered_benefits
                ),
            ),
            (
                "Analysis Mode",
                (
                    prepared_analysis[
                        "Analysis Mode"
                    ].iloc[0]
                    if (
                        not prepared_analysis.empty
                        and "Analysis Mode"
                        in prepared_analysis.columns
                    )
                    else ""
                ),
            ),
            (
                "Laser Members",
                "\n".join(lasers)
                if lasers
                else "None",
            ),
            (
                "Excluded Members",
                "\n".join(excluded)
                if excluded
                else "None",
            ),
            (
                "Notes",
                contract.get("notes", ""),
            ),
        ],
        columns=[
            "Field",
            "Value",
        ],
    )

    if (
        "Redbridge Liability"
        in prepared_analysis.columns
    ):
        large_claims = prepared_analysis[
            prepared_analysis[
                "Redbridge Liability"
            ] > 0
        ].copy()
    else:
        large_claims = prepared_analysis.iloc[
            0:0
        ].copy()

    if (
        "Coverage Status"
        in prepared_analysis.columns
    ):
        excluded_members = prepared_analysis[
            prepared_analysis[
                "Coverage Status"
            ]
            == "Excluded - No Coverage"
        ].copy()
    else:
        excluded_members = prepared_analysis.iloc[
            0:0
        ].copy()

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        summary.to_excel(
            writer,
            sheet_name="Summary",
            index=False,
        )

        large_claims.to_excel(
            writer,
            sheet_name="Large Claims",
            index=False,
        )

        prepared_analysis.to_excel(
            writer,
            sheet_name="All Members",
            index=False,
        )

        excluded_members.to_excel(
            writer,
            sheet_name="Excluded Members",
            index=False,
        )

        prepared_raw_claims.to_excel(
            writer,
            sheet_name="Raw Claims",
            index=False,
        )

        for worksheet in writer.book.worksheets:
            _style_worksheet(
                worksheet
            )

        summary_sheet = writer.book[
            "Summary"
        ]

        summary_sheet.column_dimensions[
            "A"
        ].width = 32

        summary_sheet.column_dimensions[
            "B"
        ].width = 60

        for row_number in range(
            2,
            summary_sheet.max_row + 1,
        ):
            field_name = summary_sheet.cell(
                row_number,
                1,
            ).value

            if field_name in {
                "Contract Deductible",
                "Analysis Deductible",
                "Maximum Redbridge Liability",
            }:
                summary_sheet.cell(
                    row_number,
                    2,
                ).number_format = "$#,##0.00"

    return output.getvalue()
